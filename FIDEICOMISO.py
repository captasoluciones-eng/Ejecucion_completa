#!/usr/bin/env python3
"""
Conciliacion FIDEICOMISO vs CAPTAVALE
=====================================

Cruza el reporte del FIDEICOMISO contra los reportes de CAPTAVALE y
clasifica los documentos que no se encuentran, segun las 7 reglas de
negocio.

REGLAS
------
1. BUSQUEDA        Comparar el "Numero de Documento" de cada fila del
                   REPORTE FIDEICOMISO contra el "IdentificadorUnicoFlujo"
                   de REPORTE CAPTAVALE 1 y 2.
2. NO ENCONTRADOS  Conservar las filas de ambas hojas del FIDEICOMISO cuyo
                   numero de documento no exista en ninguno de los dos
                   reportes de CAPTAVALE.
3. DIFERENCIAS     Para los registros encontrados, comparar
                   "CarteraController Total VN" contra el importe SLD
                   ("MontoDelFlujo" de CAPTAVALE).
4. POR CONTRATO    Agrupar los no encontrados por "Numero de Contrato". Las
                   decisiones siguientes consideran el conjunto completo de
                   documentos del contrato.
5. CSTI            Si todos los documentos del contrato inician con C, I, T
                   o S, existen las cuatro letras, y comparten una
                   terminacion de dos digitos para el mismo conjunto,
                   clasificar el contrato como CSTI.
6. STI             El resto de contratos cuyos documentos solo inician con
                   C, I, T, S pero no cumplen la regla 5.
7. INTEGRIDAD      La suma de documentos CSTI, STI y excluidos debe ser
                   igual al total de documentos no encontrados.

USO
---
    python FIDEICOMISO.py <carpeta> [--clave F12540] [--salida ./salida]

La carpeta debe contener:
    REPORTE CAPTA 1 <clave>.csv
    REPORTE CAPTA 2 <clave>.csv
    REPORTE FIDEICOMISO <clave>.xlsx     (tambien acepta .csv)

Ejemplo:
    python FIDEICOMISO.py "G:/Mi unidad/.../F12540" --clave F12540

SALIDA
------
    CONCILIACION <clave>.xlsx    entregable principal, cuatro hojas:
                                 Resumen, No encontrados, Diferencias
                                 importe y Contratos.
    *.csv, resumen_<clave>.json  mismos datos, para uso automatizado.

DEPENDENCIAS
------------
    pip install openpyxl
"""

import argparse
import csv
import json
import sys
from collections import defaultdict
from pathlib import Path

# --------------------------------------------------------------------------
# Configuracion
# --------------------------------------------------------------------------

# Diferencia maxima para considerar que dos importes son iguales (regla 3).
TOLERANCIA_IMPORTE = 0.01

# Nombres de columna en cada origen. Se localizan por nombre, asi que el
# orden real de las columnas puede cambiar sin romper el proceso.
COLS_CAPTA = {
    "contrato": "NumeroDeContrato",
    "documento": "IdentificadorUnicoFlujo",
    "importe": "MontoDelFlujo",
}

COLS_FIDEICOMISO = {
    "contrato": "Numero de Contrato",
    "documento": "Numero de Documento",
    "importe": "CarteraController Total VN",
    "concepto": "Concepto",
    "subtipo": "Subtipo Producto",
}

LETRAS_VALIDAS = {"C", "I", "T", "S"}


# --------------------------------------------------------------------------
# Utilidades
# --------------------------------------------------------------------------

def a_numero(valor):
    """Convierte a float tolerando None, cadenas vacias y separadores."""
    if valor is None or valor == "":
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    limpio = str(valor).replace("$", "").replace(",", "").strip()
    try:
        return float(limpio)
    except ValueError:
        return 0.0


def indices_de(encabezado, columnas):
    """Mapea {clave logica: indice} a partir de la fila de encabezado."""
    posicion = {}
    for i, nombre in enumerate(encabezado):
        if nombre is not None:
            posicion[str(nombre).strip().lstrip("\ufeff")] = i

    indices = {}
    faltantes = []
    for clave, nombre in columnas.items():
        if nombre in posicion:
            indices[clave] = posicion[nombre]
        else:
            faltantes.append(nombre)

    if faltantes:
        raise SystemExit(
            "Faltan columnas: {}\nEncabezado encontrado: {}".format(
                ", ".join(faltantes), " | ".join(str(h) for h in encabezado)
            )
        )
    return indices


# --------------------------------------------------------------------------
# Lectura de los insumos
# --------------------------------------------------------------------------

def cargar_captavale(rutas):
    """
    Regla 1: construye el indice de documentos de CAPTAVALE.

    Se mantiene UN INDICE POR ARCHIVO en lugar de uno solo combinado. Cuesta
    lo mismo, pero permite demostrar despues cuantas coincidencias aporto
    cada archivo: si CAPTA 2 aporta coincidencias, quedo probado que se leyo.

    Nota: CAPTA 1 y CAPTA 2 no son reportes independientes, son un mismo
    export partido por el limite de 1,000,000 de filas del origen.

    Devuelve (indices, estadisticas) donde indices es una lista de dicts
    {numero de documento: importe SLD}, uno por archivo.
    """
    indices = []
    estadisticas = []

    for ruta in rutas:
        ruta = Path(ruta)
        print(f"  Leyendo {ruta.name} ...", flush=True)

        indice = {}
        filas = 0
        duplicados_internos = 0

        with open(ruta, "r", encoding="utf-8-sig", newline="") as fh:
            lector = csv.reader(fh)
            try:
                encabezado = next(lector)
            except StopIteration:
                encabezado = []
            if encabezado:
                idx = indices_de(encabezado, COLS_CAPTA)
                for fila in lector:
                    if len(fila) <= idx["documento"]:
                        continue
                    documento = fila[idx["documento"]].strip()
                    if not documento:
                        continue
                    filas += 1
                    if documento in indice:
                        duplicados_internos += 1
                        continue
                    indice[documento] = a_numero(fila[idx["importe"]])

        indices.append(indice)
        estadisticas.append({
            "archivo": ruta.name,
            "tamano_mb": round(ruta.stat().st_size / 1024 / 1024, 1),
            "filas": filas,
            "documentos_unicos": len(indice),
            "duplicados_internos": duplicados_internos,
            "coincidencias": 0,   # se llena durante el cruce
        })
        print(f"    {filas:,} filas -> {len(indice):,} documentos unicos"
              + (f" ({duplicados_internos:,} duplicados)" if duplicados_internos else ""))

    # Un documento que aparezca en los dos archivos indicaria traslape entre
    # ellos, es decir que NO son un export partido limpiamente.
    traslape = 0
    if len(indices) == 2:
        chico, grande = sorted(indices, key=len)
        traslape = sum(1 for d in chico if d in grande)

    total_unicos = sum(len(d) for d in indices)
    print(f"  CAPTAVALE: {total_unicos:,} documentos unicos en total, "
          f"{traslape:,} traslapados entre archivos")

    return indices, estadisticas, traslape


def iter_fideicomiso(ruta):
    """
    Recorre el reporte del FIDEICOMISO fila por fila, sin cargarlo entero en
    memoria. Acepta .xlsx (todas las hojas, regla 2) y .csv.

    Entrega tuplas: (contrato, documento, importe, concepto, subtipo)
    """
    ruta = Path(ruta)

    if ruta.suffix.lower() == ".csv":
        yield from _iter_fideicomiso_csv(ruta)
    else:
        yield from _iter_fideicomiso_xlsx(ruta)


def _iter_fideicomiso_csv(ruta):
    with open(ruta, "r", encoding="utf-8-sig", newline="") as fh:
        lector = csv.reader(fh)
        encabezado = next(lector)
        idx = indices_de(encabezado, COLS_FIDEICOMISO)
        for fila in lector:
            if len(fila) <= idx["documento"]:
                continue
            documento = fila[idx["documento"]].strip()
            if not documento:
                continue
            yield (
                fila[idx["contrato"]].strip(),
                documento,
                a_numero(fila[idx["importe"]]),
                fila[idx["concepto"]],
                fila[idx["subtipo"]],
            )


def _iter_fideicomiso_xlsx(ruta):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit(
            "Falta la dependencia openpyxl. Instalala con:\n"
            "    pip install openpyxl"
        )

    # read_only=True activa el modo streaming: openpyxl no materializa la
    # hoja completa, que aqui son mas de un millon de filas.
    libro = load_workbook(ruta, read_only=True, data_only=True)

    try:
        for hoja in libro.worksheets:
            print(f"  Hoja '{hoja.title}' ...", flush=True)
            idx = None
            for fila in hoja.iter_rows(values_only=True):
                if fila is None:
                    continue

                # El encabezado no siempre esta en la primera fila (en estos
                # reportes viene en la segunda), asi que se busca.
                if idx is None:
                    if any(c == COLS_FIDEICOMISO["documento"] for c in fila if c):
                        idx = indices_de(list(fila), COLS_FIDEICOMISO)
                    continue

                if len(fila) <= idx["documento"]:
                    continue
                documento = fila[idx["documento"]]
                if documento is None or str(documento).strip() == "":
                    continue

                yield (
                    str(fila[idx["contrato"]] or "").strip(),
                    str(documento).strip(),
                    a_numero(fila[idx["importe"]]),
                    fila[idx["concepto"]] or "",
                    fila[idx["subtipo"]] or "",
                )

            if idx is None:
                print(f"    (sin encabezado reconocible, hoja omitida)")
    finally:
        libro.close()


def conciliar_inverso(rutas, documentos_buscados):
    """
    Conciliacion en sentido inverso: CAPTAVALE contra FIDEICOMISO.

    Recibe los documentos de CAPTAVALE que no aparecieron en el fideicomiso y
    recupera su fila completa con una segunda lectura de los CSV. Se hace en
    una segunda pasada, y no guardando todo durante la primera, porque el
    conjunto resultante es diminuto frente al millon de registros: cuesta unos
    segundos de lectura y evita cargar el contrato de cada documento en memoria.

    Devuelve los mismos agrupamientos que la conciliacion directa, para que las
    reglas 4, 5, 6 y 7 se apliquen igual en esta direccion.
    """
    if not documentos_buscados:
        return {}

    por_contrato = defaultdict(list)

    for ruta in rutas:
        ruta = Path(ruta)
        with open(ruta, "r", encoding="utf-8-sig", newline="") as fh:
            lector = csv.reader(fh)
            try:
                encabezado = next(lector)
            except StopIteration:
                continue
            idx = indices_de(encabezado, COLS_CAPTA)

            for fila in lector:
                if len(fila) <= idx["documento"]:
                    continue
                documento = fila[idx["documento"]].strip()
                if documento not in documentos_buscados:
                    continue
                contrato = fila[idx["contrato"]].strip()
                por_contrato[contrato].append({
                    "documento": documento,
                    "letra": documento[0].upper() if documento else "",
                    "terminacion": documento[-2:],
                    "concepto": "",
                    "subtipo": "",
                    "importe": a_numero(fila[idx["importe"]]),
                    "origen": ruta.name,
                })

    return por_contrato


# --------------------------------------------------------------------------
# Reglas de clasificacion
# --------------------------------------------------------------------------

def clasificar_contrato(documentos):
    """
    Reglas 5 y 6. Recibe todos los documentos NO ENCONTRADOS de un contrato.

    CSTI     todos inician con C/I/T/S, estan las cuatro letras, y cada
             conjunto que comparte terminacion de dos digitos esta completo
             con las cuatro.
    STI      resto de contratos que solo usan C/I/T/S.
    EXCLUIDO algun documento inicia con otra letra.

    Criterio adoptado: se exige que TODOS los conjuntos por terminacion esten
    completos ("todos los documentos del contrato"). Si existe un conjunto
    completo pero sobran documentos sueltos con otra terminacion, el contrato
    cae en STI.
    """
    letras = {d["letra"] for d in documentos}
    if not letras or not letras <= LETRAS_VALIDAS:
        return "EXCLUIDO"

    por_terminacion = defaultdict(set)
    for d in documentos:
        por_terminacion[d["terminacion"]].add(d["letra"])

    todos_completos = all(
        conjunto >= LETRAS_VALIDAS for conjunto in por_terminacion.values()
    )
    estan_las_cuatro = letras >= LETRAS_VALIDAS

    return "CSTI" if (todos_completos and estan_las_cuatro) else "STI"


# --------------------------------------------------------------------------
# Proceso principal
# --------------------------------------------------------------------------

def conciliar(indices, estadisticas, ruta_fideicomiso):
    """
    Aplica las reglas 1 a 4 recorriendo el FIDEICOMISO una sola vez.

    Ademas de conciliar, deja rastro de evidencia: cuantas coincidencias
    aporto cada archivo de CAPTAVALE y que documentos de CAPTAVALE nunca
    fueron usados. Eso permite verificar que ambos archivos se leyeron y que
    el resultado es consistente en las dos direcciones.
    """
    total = 0
    encontrados = 0
    diferencias = []
    no_encontrados_por_contrato = defaultdict(list)

    # Documentos de CAPTAVALE efectivamente usados, por archivo.
    usados = [set() for _ in indices]

    for contrato, documento, importe, concepto, subtipo in iter_fideicomiso(ruta_fideicomiso):
        total += 1
        if total % 250_000 == 0:
            print(f"    {total:,} filas procesadas ...", flush=True)

        # Regla 1: se busca en cada archivo por separado para saber cual
        # de los dos aporto la coincidencia.
        origen = None
        importe_sld = None
        for i, indice in enumerate(indices):
            if documento in indice:
                origen = i
                importe_sld = indice[documento]
                break

        if origen is not None:
            encontrados += 1
            estadisticas[origen]["coincidencias"] += 1
            usados[origen].add(documento)

            diferencia = abs(importe - importe_sld)
            if diferencia > TOLERANCIA_IMPORTE:  # Regla 3
                diferencias.append({
                    "contrato": contrato,
                    "documento": documento,
                    "cartera_vn": importe,
                    "importe_sld": importe_sld,
                    "diferencia": round(diferencia, 4),
                    "origen": estadisticas[origen]["archivo"],
                })
        else:                                   # Regla 2
            no_encontrados_por_contrato[contrato].append({  # Regla 4
                "documento": documento,
                "letra": documento[0].upper() if documento else "",
                "terminacion": documento[-2:],
                "concepto": concepto,
                "subtipo": subtipo,
                "importe": importe,
            })

    # Verificaciones cruzadas
    for i, st in enumerate(estadisticas):
        st["documentos_usados"] = len(usados[i])
        st["documentos_sin_usar"] = st["documentos_unicos"] - len(usados[i])

    # Documentos de CAPTAVALE que nunca aparecieron en el FIDEICOMISO.
    # Se construye por comprension para no materializar la diferencia de
    # conjuntos completa: el resultado es pequeno aunque los indices no lo sean.
    sin_contraparte = {
        documento
        for i, indice in enumerate(indices)
        for documento in indice
        if documento not in usados[i]
    }

    distintos_usados = sum(len(u) for u in usados)
    verificacion = {
        # Si el FIDEICOMISO trae el mismo documento mas de una vez, hay mas
        # coincidencias que documentos distintos de CAPTAVALE utilizados.
        "documentos_repetidos_fideicomiso": encontrados - distintos_usados,
        # Documentos que estan en CAPTAVALE pero no aparecen en el
        # FIDEICOMISO: el cruce en sentido inverso.
        "captavale_sin_contraparte": len(sin_contraparte),
    }

    return (total, encontrados, diferencias, no_encontrados_por_contrato,
            verificacion, sin_contraparte)


def escribir_csv(ruta, columnas, filas):
    with open(ruta, "w", encoding="utf-8-sig", newline="") as fh:
        escritor = csv.DictWriter(fh, fieldnames=columnas)
        escritor.writeheader()
        escritor.writerows(filas)


# --------------------------------------------------------------------------
# Salida en Excel
# --------------------------------------------------------------------------

# Ancho de columna y formato numerico por nombre de campo.
ANCHOS = {
    "contrato": 18, "documento": 20, "letra": 8, "terminacion": 13,
    "concepto": 12, "subtipo": 20, "cartera_vn": 16, "importe_sld": 18,
    "diferencia": 14, "clasificacion_contrato": 22, "documentos": 14,
    "clasificacion": 16,
}
FORMATO_IMPORTE = {"cartera_vn", "importe_sld", "diferencia"}

COLS_DETALLE_XLSX = ["contrato", "documento", "letra", "terminacion", "concepto",
                     "subtipo", "cartera_vn", "clasificacion_contrato"]

TITULOS = {
    "contrato": "Numero de Contrato",
    "documento": "Numero de Documento",
    "letra": "Letra",
    "terminacion": "Terminacion",
    "concepto": "Concepto",
    "subtipo": "Subtipo Producto",
    "cartera_vn": "CarteraController Total VN",
    "importe_sld": "Importe SLD (CAPTAVALE)",
    "diferencia": "Diferencia",
    "clasificacion_contrato": "Clasificacion del contrato",
    "documentos": "Documentos no encontrados",
    "clasificacion": "Clasificacion",
    "origen": "Archivo CAPTAVALE de origen",
}


def _hoja_de_datos(libro, titulo, columnas, filas):
    """Agrega una hoja con encabezado fijo, filtro y formato de importes."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    hoja = libro.create_sheet(titulo)

    encabezado = [TITULOS.get(c, c) for c in columnas]
    hoja.append(encabezado)
    for celda in hoja[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="1F4E5F")
        celda.alignment = Alignment(vertical="center")
    hoja.row_dimensions[1].height = 22

    for fila in filas:
        hoja.append([fila.get(c, "") for c in columnas])

    for i, clave in enumerate(columnas, start=1):
        letra = get_column_letter(i)
        hoja.column_dimensions[letra].width = ANCHOS.get(clave, 16)
        if clave in FORMATO_IMPORTE:
            for celda in hoja[letra][1:]:
                celda.number_format = "#,##0.00"

    hoja.freeze_panes = "A2"
    if filas:
        hoja.auto_filter.ref = f"A1:{get_column_letter(len(columnas))}{len(filas) + 1}"
    return hoja


def escribir_excel(ruta, resumen, detalle_no_enc, por_clase, diferencias,
                   detalle_contratos, inv_detalle, inv_contratos):
    """
    Genera el archivo .xlsx entregable.

    Hojas:
        Resumen                 cifras principales y control de integridad
        Verificacion            evidencia de que ambos archivos se leyeron
        No encontrados          los faltantes completos, sin separar
        CSTI (regla 5)          documentos de contratos clasificados CSTI
        STI (regla 6)           documentos de contratos clasificados STI
        Excluidos               solo si existen
        Diferencias importe     conciliados con monto distinto
        Contratos               un renglon por contrato
        Inverso - faltantes     documentos de CAPTAVALE que no estan en el
                                fideicomiso (cruce en sentido contrario)
        Inverso - contratos     esos mismos, agrupados por contrato
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    libro = Workbook()
    libro.remove(libro.active)          # quita la hoja vacia por defecto

    titulo_font = Font(bold=True, size=15, color="1F4E5F")
    seccion_fill = PatternFill("solid", fgColor="EDF1F2")

    # ------------------------------------------------------------------
    # Hoja Resumen
    # ------------------------------------------------------------------
    hoja = libro.create_sheet("Resumen")
    hoja.column_dimensions["A"].width = 44
    hoja.column_dimensions["B"].width = 20
    hoja.sheet_view.showGridLines = False

    hoja["A1"] = f"CONCILIACION {resumen['clave']}"
    hoja["A1"].font = titulo_font
    hoja["A2"] = "FIDEICOMISO contra CAPTAVALE"
    hoja["A2"].font = Font(italic=True, color="666666")

    renglones = [
        ("", ""),
        ("Documentos revisados", resumen["filas_fideicomiso"]),
        ("", ""),
        ("Conciliados", resumen["encontrados"]),
        ("   con diferencia de importe", resumen["diferencias_importe"]),
        ("", ""),
        ("Sin contraparte", resumen["no_encontrados"]),
        ("   contratos afectados", resumen["contratos_no_encontrados"]),
        ("", ""),
        ("CSTI (regla 5)", resumen["documentos_csti"]),
        ("   en contratos", resumen["contratos_csti"]),
        ("STI (regla 6)", resumen["documentos_sti"]),
        ("   en contratos", resumen["contratos_sti"]),
        ("Excluidos", resumen["documentos_excluidos"]),
        ("", ""),
        ("Control de integridad (regla 7)", resumen["control_integridad"]),
        ("", ""),
        ("CRUCE INVERSO (CAPTAVALE contra FIDEICOMISO)", ""),
        ("Documentos de CAPTAVALE sin contraparte", resumen["inverso"]["documentos"]),
        ("   contratos afectados", resumen["inverso"]["contratos"]),
        ("   CSTI (regla 5)", resumen["inverso"]["csti"]),
        ("   STI (regla 6)", resumen["inverso"]["sti"]),
        ("   excluidos", resumen["inverso"]["excluidos"]),
        ("Control de integridad del inverso", resumen["inverso"]["control_integridad"]),
    ]

    fila_actual = 3
    filas_control = []          # (fila, cuadra) de cada control de integridad
    for etiqueta, valor in renglones:
        hoja.cell(row=fila_actual, column=1, value=etiqueta)
        if valor != "":
            celda = hoja.cell(row=fila_actual, column=2, value=valor)
            if isinstance(valor, int):
                celda.number_format = "#,##0"
                celda.alignment = Alignment(horizontal="right")
        if etiqueta and not etiqueta.startswith("   "):
            hoja.cell(row=fila_actual, column=1).font = Font(bold=True)
            hoja.cell(row=fila_actual, column=2).font = Font(bold=True)
        if etiqueta.startswith("Control de integridad"):
            filas_control.append((fila_actual, valor == "OK"))
        if etiqueta.startswith("CRUCE INVERSO"):
            hoja.cell(row=fila_actual, column=1).fill = seccion_fill
        fila_actual += 1

    # Resalta los controles de integridad: es lo primero que hay que mirar.
    for fila, cuadra in filas_control:
        for col in (1, 2):
            celda = hoja.cell(row=fila, column=col)
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = PatternFill("solid", fgColor="1B6E45" if cuadra else "B03A2E")

    # ------------------------------------------------------------------
    # Hoja Verificacion: evidencia de que ambos archivos se leyeron
    # ------------------------------------------------------------------
    ver = libro.create_sheet("Verificacion")
    ver.sheet_view.showGridLines = False
    for col, ancho in zip("ABCDEF", (34, 14, 16, 20, 18, 18)):
        ver.column_dimensions[col].width = ancho

    ver["A1"] = "VERIFICACION DE LECTURA"
    ver["A1"].font = titulo_font
    ver["A2"] = ("Cada archivo de CAPTAVALE se indexa por separado, de modo que "
                 "se puede comprobar cuantas coincidencias aporto cada uno.")
    ver["A2"].font = Font(italic=True, color="666666")

    encabezado = ["Archivo", "Tamano MB", "Filas leidas", "Documentos unicos",
                  "Coincidencias", "Sin usar"]
    ver.append([])
    ver.append(encabezado)
    fila_enc = ver.max_row
    for celda in ver[fila_enc]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="1F4E5F")

    for st in resumen["archivos_captavale"]:
        ver.append([st["archivo"], st["tamano_mb"], st["filas"],
                    st["documentos_unicos"], st["coincidencias"],
                    st.get("documentos_sin_usar", 0)])
        for c in ver[ver.max_row][1:]:
            c.number_format = "#,##0"

    # Marca en rojo cualquier archivo que no haya aportado coincidencias:
    # seria la senal de que no se leyo o de que no corresponde.
    for i, st in enumerate(resumen["archivos_captavale"]):
        if st["coincidencias"] == 0:
            for c in ver[fila_enc + 1 + i]:
                c.fill = PatternFill("solid", fgColor="F5D5D0")

    ver.append([])
    ver.append(["COMPROBACIONES CRUZADAS"])
    ver[ver.max_row][0].font = Font(bold=True)
    ver[ver.max_row][0].fill = seccion_fill

    comprobaciones = [
        ("Documentos traslapados entre CAPTA 1 y CAPTA 2",
         resumen["traslape_entre_archivos"],
         "Debe ser 0: los archivos son un mismo export partido, no se repiten."),
        ("Documentos repetidos dentro del FIDEICOMISO",
         resumen["documentos_repetidos_fideicomiso"],
         "Filas del fideicomiso que apuntan a un documento ya contado."),
        ("Documentos de CAPTAVALE sin contraparte",
         resumen["captavale_sin_contraparte"],
         "Cruce inverso: estan en CAPTAVALE pero no en el fideicomiso."),
    ]
    for etiqueta, valor, nota in comprobaciones:
        ver.append([etiqueta, valor, "", nota])
        fila = ver[ver.max_row]
        fila[1].number_format = "#,##0"
        fila[1].font = Font(bold=True)
        fila[3].font = Font(color="666666", size=9)

    ver.append([])
    ver.append(["CUADRE DE TOTALES"])
    ver[ver.max_row][0].font = Font(bold=True)
    ver[ver.max_row][0].fill = seccion_fill

    total_unicos = sum(st["documentos_unicos"] for st in resumen["archivos_captavale"])
    total_coincidencias = sum(st["coincidencias"] for st in resumen["archivos_captavale"])
    cuadres = [
        ("Conciliados + sin contraparte", resumen["encontrados"] + resumen["no_encontrados"]),
        ("   debe igualar documentos revisados", resumen["filas_fideicomiso"]),
        ("Coincidencias sumadas de ambos archivos", total_coincidencias),
        ("   debe igualar conciliados", resumen["encontrados"]),
        ("Documentos unicos de CAPTAVALE", total_unicos),
    ]
    for etiqueta, valor in cuadres:
        ver.append([etiqueta, valor])
        ver[ver.max_row][1].number_format = "#,##0"
        if not etiqueta.startswith("   "):
            ver[ver.max_row][0].font = Font(bold=True)

    # ------------------------------------------------------------------
    # Hojas de detalle
    # ------------------------------------------------------------------
    _hoja_de_datos(libro, "No encontrados", COLS_DETALLE_XLSX, detalle_no_enc)
    _hoja_de_datos(libro, "CSTI (regla 5)", COLS_DETALLE_XLSX, por_clase["CSTI"])
    _hoja_de_datos(libro, "STI (regla 6)", COLS_DETALLE_XLSX, por_clase["STI"])
    if por_clase["EXCLUIDO"]:
        _hoja_de_datos(libro, "Excluidos", COLS_DETALLE_XLSX, por_clase["EXCLUIDO"])

    _hoja_de_datos(libro, "Diferencias importe",
                   ["contrato", "documento", "cartera_vn", "importe_sld",
                    "diferencia", "origen"],
                   diferencias)
    _hoja_de_datos(libro, "Contratos",
                   ["contrato", "documentos", "clasificacion"],
                   detalle_contratos)

    # ------------------------------------------------------------------
    # Cruce inverso: CAPTAVALE contra FIDEICOMISO
    # ------------------------------------------------------------------
    _hoja_de_datos(libro, "Inverso - faltantes",
                   ["contrato", "documento", "letra", "terminacion",
                    "importe_sld", "origen", "clasificacion_contrato"],
                   inv_detalle)
    _hoja_de_datos(libro, "Inverso - contratos",
                   ["contrato", "documentos", "clasificacion"],
                   inv_contratos)

    libro.save(ruta)


def main():
    parser = argparse.ArgumentParser(
        description="Conciliacion FIDEICOMISO vs CAPTAVALE",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("carpeta", help="Carpeta que contiene los tres archivos")
    parser.add_argument("--clave", default="F12540",
                        help="Clave del fideicomiso en los nombres de archivo (default: F12540)")
    parser.add_argument("--salida", default="salida",
                        help="Carpeta donde escribir los resultados (default: ./salida)")
    args = parser.parse_args()

    carpeta = Path(args.carpeta)
    if not carpeta.is_dir():
        raise SystemExit(f"No existe la carpeta: {carpeta}")

    capta = [
        carpeta / f"REPORTE CAPTA 1 {args.clave}.csv",
        carpeta / f"REPORTE CAPTA 2 {args.clave}.csv",
    ]

    # El FIDEICOMISO puede venir como .xlsx o .csv.
    fideicomiso = carpeta / f"REPORTE FIDEICOMISO {args.clave}.xlsx"
    if not fideicomiso.exists():
        alterno = carpeta / f"REPORTE FIDEICOMISO {args.clave}.csv"
        if alterno.exists():
            fideicomiso = alterno

    for ruta in capta + [fideicomiso]:
        if not ruta.exists():
            raise SystemExit(f"No se encontro el archivo: {ruta}")

    print(f"\nCONCILIACION {args.clave}")
    print("=" * 60)

    print("\n[1/3] Cargando CAPTAVALE ...")
    indices, estadisticas, traslape = cargar_captavale(capta)

    print(f"\n[2/4] Recorriendo {fideicomiso.name} ...")
    total, encontrados, diferencias, no_enc, verificacion, sin_contraparte = conciliar(
        indices, estadisticas, fideicomiso)

    print("\n[3/4] Clasificando contratos no encontrados ...")
    # Reglas 5 y 6 en listas separadas: cada una va a su propia hoja.
    por_clase = {"CSTI": [], "STI": [], "EXCLUIDO": []}
    detalle_no_enc = []
    detalle_contratos = []
    conteo = {"CSTI": 0, "STI": 0, "EXCLUIDO": 0}
    contratos_por_clase = {"CSTI": 0, "STI": 0, "EXCLUIDO": 0}
    no_encontrados = 0

    for contrato, documentos in no_enc.items():
        clase = clasificar_contrato(documentos)          # Reglas 5 y 6
        conteo[clase] += len(documentos)
        contratos_por_clase[clase] += 1
        no_encontrados += len(documentos)

        detalle_contratos.append({
            "contrato": contrato,
            "documentos": len(documentos),
            "clasificacion": clase,
        })
        for d in documentos:
            registro = {
                "contrato": contrato,
                "documento": d["documento"],
                "letra": d["letra"],
                "terminacion": d["terminacion"],
                "concepto": d["concepto"],
                "subtipo": d["subtipo"],
                "cartera_vn": d["importe"],
                "clasificacion_contrato": clase,
            }
            detalle_no_enc.append(registro)
            por_clase[clase].append(registro)

    # Regla 7: control de integridad
    suma = conteo["CSTI"] + conteo["STI"] + conteo["EXCLUIDO"]
    cuadra = (suma == no_encontrados)

    # ------------------------------------------------------------------
    # Conciliacion inversa: CAPTAVALE contra FIDEICOMISO
    # ------------------------------------------------------------------
    print(f"\n[4/4] Conciliacion inversa ({len(sin_contraparte):,} documentos) ...")
    inv_por_contrato = conciliar_inverso(capta, sin_contraparte)

    inv_detalle = []
    inv_contratos = []
    inv_conteo = {"CSTI": 0, "STI": 0, "EXCLUIDO": 0}

    for contrato, documentos in inv_por_contrato.items():
        clase = clasificar_contrato(documentos)          # Reglas 5 y 6
        inv_conteo[clase] += len(documentos)
        inv_contratos.append({
            "contrato": contrato,
            "documentos": len(documentos),
            "clasificacion": clase,
        })
        for d in documentos:
            inv_detalle.append({
                "contrato": contrato,
                "documento": d["documento"],
                "letra": d["letra"],
                "terminacion": d["terminacion"],
                "importe_sld": d["importe"],
                "origen": d["origen"],
                "clasificacion_contrato": clase,
            })

    inv_total = len(inv_detalle)
    inv_suma = inv_conteo["CSTI"] + inv_conteo["STI"] + inv_conteo["EXCLUIDO"]
    inv_cuadra = (inv_suma == inv_total == len(sin_contraparte))

    inverso = {
        "documentos": inv_total,
        "contratos": len(inv_por_contrato),
        "csti": inv_conteo["CSTI"],
        "sti": inv_conteo["STI"],
        "excluidos": inv_conteo["EXCLUIDO"],
        "control_integridad": "OK" if inv_cuadra else "DESCUADRE",
    }

    resumen = {
        "clave": args.clave,
        "filas_fideicomiso": total,
        "encontrados": encontrados,
        "diferencias_importe": len(diferencias),
        "no_encontrados": no_encontrados,
        "contratos_no_encontrados": len(no_enc),
        "documentos_csti": conteo["CSTI"],
        "documentos_sti": conteo["STI"],
        "documentos_excluidos": conteo["EXCLUIDO"],
        "contratos_csti": contratos_por_clase["CSTI"],
        "contratos_sti": contratos_por_clase["STI"],
        "contratos_excluidos": contratos_por_clase["EXCLUIDO"],
        "control_integridad": "OK" if cuadra else "DESCUADRE",
        "archivos_captavale": estadisticas,
        "traslape_entre_archivos": traslape,
        "documentos_repetidos_fideicomiso": verificacion["documentos_repetidos_fideicomiso"],
        "captavale_sin_contraparte": verificacion["captavale_sin_contraparte"],
        "inverso": inverso,
    }

    salida = Path(args.salida)
    salida.mkdir(parents=True, exist_ok=True)

    COLS_DETALLE = ["contrato", "documento", "letra", "terminacion", "concepto",
                    "subtipo", "cartera_vn", "clasificacion_contrato"]

    escribir_csv(salida / f"no_encontrados_{args.clave}.csv", COLS_DETALLE, detalle_no_enc)
    escribir_csv(salida / f"csti_{args.clave}.csv", COLS_DETALLE, por_clase["CSTI"])
    escribir_csv(salida / f"sti_{args.clave}.csv", COLS_DETALLE, por_clase["STI"])
    escribir_csv(salida / f"inverso_no_encontrados_{args.clave}.csv",
                 ["contrato", "documento", "letra", "terminacion", "importe_sld",
                  "origen", "clasificacion_contrato"],
                 inv_detalle)
    escribir_csv(salida / f"diferencias_importe_{args.clave}.csv",
                 ["contrato", "documento", "cartera_vn", "importe_sld", "diferencia", "origen"],
                 diferencias)
    escribir_csv(salida / f"contratos_{args.clave}.csv",
                 ["contrato", "documentos", "clasificacion"],
                 detalle_contratos)
    with open(salida / f"resumen_{args.clave}.json", "w", encoding="utf-8") as fh:
        json.dump(resumen, fh, indent=2, ensure_ascii=False)

    # Entregable principal: un solo Excel, con las reglas 5 y 6 separadas.
    ruta_excel = salida / f"CONCILIACION {args.clave}.xlsx"
    escribir_excel(ruta_excel, resumen, detalle_no_enc, por_clase, diferencias,
                   detalle_contratos, inv_detalle, inv_contratos)

    print("\n" + "=" * 60)
    print(f"RESUMEN {args.clave}")
    print("=" * 60)
    print(f"Filas del FIDEICOMISO            {total:>12,}")
    print(f"  Encontrados                    {encontrados:>12,}")
    print(f"    con diferencia de importe    {len(diferencias):>12,}")
    print(f"  No encontrados                 {no_encontrados:>12,}")
    print(f"    contratos afectados          {len(no_enc):>12,}")
    print(f"    documentos CSTI              {conteo['CSTI']:>12,}")
    print(f"    documentos STI               {conteo['STI']:>12,}")
    print(f"    documentos excluidos         {conteo['EXCLUIDO']:>12,}")
    print("-" * 60)
    print(f"Regla 7 (integridad): {suma:,} == {no_encontrados:,} -> "
          f"{'OK' if cuadra else 'DESCUADRE'}")

    print("-" * 60)
    print("VERIFICACION DE LECTURA")
    for st in estadisticas:
        marca = "OK " if st["coincidencias"] > 0 else "!! "
        print(f"  {marca}{st['archivo']}")
        print(f"       {st['filas']:>10,} filas   "
              f"{st['documentos_unicos']:>10,} unicos   "
              f"{st['coincidencias']:>10,} coincidencias")
    print(f"  Traslape entre archivos          {traslape:>12,}  (esperado 0)")
    print(f"  Repetidos en el FIDEICOMISO      "
          f"{verificacion['documentos_repetidos_fideicomiso']:>12,}")
    print(f"  CAPTAVALE sin contraparte        "
          f"{verificacion['captavale_sin_contraparte']:>12,}")

    print("-" * 60)
    print("CRUCE INVERSO (CAPTAVALE contra FIDEICOMISO)")
    print(f"  Documentos sin contraparte       {inverso['documentos']:>12,}")
    print(f"    contratos afectados            {inverso['contratos']:>12,}")
    print(f"    CSTI                           {inverso['csti']:>12,}")
    print(f"    STI                            {inverso['sti']:>12,}")
    print(f"    excluidos                      {inverso['excluidos']:>12,}")
    print(f"  Integridad del inverso: {inv_suma:,} == {inv_total:,} -> "
          f"{inverso['control_integridad']}")
    print("=" * 60)
    print(f"\nEXCEL: {ruta_excel.resolve()}")
    print(f"Detalle en CSV y resumen JSON: {salida.resolve()}")

    if not cuadra:
        sys.exit(1)


if __name__ == "__main__":
    main()
